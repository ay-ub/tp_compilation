import ply.yacc as yacc
from lexer import tokens
import openpyxl
from datetime import datetime
from quad_generator import QuadGenerator

# Initialize quad generator
quad_gen = QuadGenerator()

# Load values from an Excel file
def load_excel_values(file_path, sheet_name="Sheet1"):
    cell_values = {}
    workbook = openpyxl.load_workbook(file_path, data_only=True)  # Read calculated values
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None:
                cell_values[cell.coordinate] = cell.value  # Ex: {"A1": 10, "B2": 5}
    
    return cell_values

cell_values = load_excel_values("data.xlsx")

precedence = (
    ('left', 'PLUS', 'MINUS'),
    ('left', 'TIMES', 'DIVIDE'),
    ('right', 'POWER')
)

def p_expression_binop(p):
    """expression : expression PLUS expression
                  | expression MINUS expression
                  | expression TIMES expression
                  | expression DIVIDE expression
                  | expression POWER expression"""
    temp = quad_gen.new_temp()
    if p[2] == '+':
        p[0] = p[1] + p[3]
        quad_gen.emit('+', str(p[1]), str(p[3]), temp)
    elif p[2] == '-':
        p[0] = p[1] - p[3]
        quad_gen.emit('-', str(p[1]), str(p[3]), temp)
    elif p[2] == '*':
        p[0] = p[1] * p[3]
        quad_gen.emit('*', str(p[1]), str(p[3]), temp)
    elif p[2] == '/':
        if p[3] != 0:
            p[0] = p[1] / p[3]
            quad_gen.emit('/', str(p[1]), str(p[3]), temp)
        else:
            print("error: Cannot divide by zero")
    elif p[2] == '^':
        p[0] = p[1] ** p[3]
        quad_gen.emit('^', str(p[1]), str(p[3]), temp)

def p_expression_parens(p):
    "expression : LPAREN expression RPAREN"
    p[0] = p[2]

def p_expression_number(p):
    "expression : NUMBER"
    temp = quad_gen.new_temp()
    quad_gen.emit('=', str(p[1]), '_', temp)
    p[0] = p[1]

def p_expression_string(p):
    "expression : STRING"
    p[0] = p[1]

# Read an Excel cell or range
def p_expression_cell_or_range(p):
    """expression : cell_ref 
                 | cell_range"""
    p[0] = p[1]

def p_cell_ref(p):
    "cell_ref : CELL_REF"
    cell = p[1].upper()
    value = cell_values.get(cell, 0)
    temp = quad_gen.new_temp()
    quad_gen.emit('load', cell, '_', temp)
    p[0] = value

def p_cell_range(p):
    "cell_range : CELL_REF COLON CELL_REF"
    start, end = p[1].upper(), p[3].upper()
    start_col, start_row = start[0], int(start[1:])
    end_col, end_row = end[0], int(end[1:])
    
    values = []
    if start_col == end_col:  # Single column
        for i in range(start_row, end_row + 1):
            cell_value = cell_values.get(f"{start_col}{i}", 0)
            # Handle nested function results
            if isinstance(cell_value, list):
                values.extend(cell_value)
            else:
                values.append(cell_value)
        p[0] = values
    else:
        print("Multi-column ranges are not supported yet.")
        p[0] = []

# Handle Excel functions

def p_expression_function(p):
    """expression : SUM LPAREN arguments RPAREN
                  | AVERAGE LPAREN arguments RPAREN
                  | COUNT LPAREN arguments RPAREN
                  | MAX LPAREN arguments RPAREN
                  | MIN LPAREN arguments RPAREN
                  | UNIQUE LPAREN arguments RPAREN
                  | TODAY LPAREN RPAREN
                  | NOW LPAREN RPAREN
                  | YEAR LPAREN expression RPAREN
                  | MONTH LPAREN expression RPAREN
                  | DAY LPAREN expression RPAREN
                  | CONCATENATE LPAREN arguments RPAREN
                  | LEFT LPAREN expression COMMA NUMBER RPAREN
                  | RIGHT LPAREN expression COMMA NUMBER RPAREN
                  | MID LPAREN expression COMMA NUMBER COMMA NUMBER RPAREN
                  | LEN LPAREN expression RPAREN
                  | LOWER LPAREN expression RPAREN
                  | UPPER LPAREN expression RPAREN
                  | TRIM LPAREN expression RPAREN
                  | FIND LPAREN expression COMMA expression COMMA NUMBER RPAREN
                  | SEARCH LPAREN expression COMMA expression COMMA NUMBER RPAREN
                  | REPLACE LPAREN expression COMMA NUMBER COMMA NUMBER COMMA expression RPAREN
                  | SUBSTITUTE LPAREN expression COMMA expression COMMA expression COMMA NUMBER RPAREN
                  | TEXT LPAREN expression COMMA STRING RPAREN
                  | VALUE LPAREN expression RPAREN
                  | PROPER LPAREN expression RPAREN
                  | REPT LPAREN expression COMMA NUMBER RPAREN
                  | EXACT LPAREN expression COMMA expression RPAREN
                  | CHAR LPAREN NUMBER RPAREN
                  | CODE LPAREN expression RPAREN"""
    func = p[1].lower()
    if func in ["sum", "average", "count", "max", "min", "unique"]:
        values = p[3]  # List of values
        if func == "sum":
            p[0] = sum(values)
        elif func == "average":
            p[0] = sum(values) / len(values) if values else 0
        elif func == "count":
            p[0] = len(values)
        elif func == "max":
            p[0] = max(values) if values else None
        elif func == "min":
            p[0] = min(values) if values else None
        elif func == "unique":
            p[0] = list(set(values))
    elif func == "today":
        p[0] = datetime.today().date()
    elif func == "now":
        p[0] = datetime.now()
    elif func in ["year", "month", "day"]:
        if isinstance(p[3], str):
            try:
                date_obj = datetime.strptime(p[3], "%Y-%m-%d")
            except ValueError:
                print("Error: Invalid date format. Use YYYY-MM-DD.")
                p[0] = None
                return
        elif isinstance(p[3], datetime):
            date_obj = p[3]
        else:
            print("Error: Invalid date input.")
            p[0] = None
            return

        if func == "year":
            p[0] = date_obj.year
        elif func == "month":
            p[0] = date_obj.month
        elif func == "day":
            p[0] = date_obj.day
    elif func == "concatenate":
        p[0] = ''.join(str(arg) for arg in p[3])
    elif func == "left":
        p[0] = p[3][:int(p[5])]
    elif func == "right":
        p[0] = p[3][-int(p[5]):]
    elif func == "mid":
        p[0] = p[3][int(p[5])-1:int(p[5])-1+int(p[7])]
    elif func == "len":
        p[0] = len(p[3])
    elif func == "lower":
        p[0] = p[3].lower()
    elif func == "upper":
        p[0] = p[3].upper()
    elif func == "trim":
        p[0] = p[3].strip()
    elif func == "find":
        # FIND is case-sensitive
        find_text = p[3]
        within_text = p[5]
        start_num = int(p[7]) - 1  # Convert to 0-based index
        pos = within_text.find(find_text, start_num)
        p[0] = pos + 1 if pos != -1 else None  # Return 1-based index
    elif func == "search":
        # SEARCH is case-insensitive
        find_text = p[3].lower()
        within_text = p[5].lower()
        start_num = int(p[7]) - 1  # Convert to 0-based index
        pos = within_text.find(find_text, start_num)
        p[0] = pos + 1 if pos != -1 else None  # Return 1-based index
    elif func == "replace":
        p[0] = p[3][:int(p[5])-1] + p[9] + p[3][int(p[5])-1+int(p[7]):]
    elif func == "substitute":
        text, old_text, new_text = p[3], p[5], p[7]
        instance_num = p[9] if len(p) > 10 else None
        if instance_num:
            parts = text.split(old_text)
            if len(parts) > instance_num:
                p[0] = old_text.join(parts[:instance_num]) + new_text + old_text.join(parts[instance_num:])
            else:
                p[0] = text
        else:
            p[0] = text.replace(old_text, new_text)
    elif func == "text":
        p[0] = format(p[3], p[5])
    elif func == "value":
        try:
            p[0] = float(p[3])
        except ValueError:
            p[0] = None
    elif func == "proper":
        p[0] = p[3].title()
    elif func == "rept":
        p[0] = p[3] * int(p[5])
    elif func == "exact":
        p[0] = p[3] == p[5]
    elif func == "char":
        p[0] = chr(int(p[3]))
    elif func == "code":
        p[0] = ord(p[3][0]) if p[3] else None

# Function arguments handling with better nested function support
def p_arguments(p):
    """arguments : arguments COMMA expression
                | expression"""
    if len(p) == 4:
        if isinstance(p[1], list):
            if isinstance(p[3], list):
                # Flatten nested function results
                p[0] = p[1] + p[3]
            else:
                p[0] = [p[1], p[3]]
        else:
            p[0] = [p[1], p[3]]
    else:
        if isinstance(p[1], list):
            p[0] = p[1]
        else:
            p[0] = [p[1]]

# Handle syntax errors
def p_error(p):
    print("Syntax error!")

# Build the parser
parser = yacc.yacc()

# Interactive testing
if __name__ == "__main__":
    while True:
        print("""--- Menu ---
              1. Display supported functions
              2. Enter an Excel expression
              3. Exit""")
        choice = int(input("Enter your choice: "))
        
        if choice == 1:
            print("""Supported functions:
                --- Arithmetic Functions ---
                                
                1) SUM(range): Calculates the sum of values in a range.
                2) AVERAGE(range): Computes the average of values in a range.
                3) COUNT(range): Counts the number of values in a range.
                4) MAX(range): Finds the maximum value in a range.
                5) MIN(range): Finds the minimum value in a range.
                6) UNIQUE(range): Extracts unique values from a range.

                --- Date and Time Functions ---
                                
                7) TODAY(): Returns today's date.
                8) NOW(): Returns the current date and time.
                9) YEAR(date): Extracts the year from a given date.
                10) MONTH(date): Extracts the month from a given date.
                11) DAY(date): Extracts the day from a given date.

                --- Text Functions ---
                                
                12) CONCATENATE(text1, text2, ...): Concatenates text strings.
                13) LEFT(text, num_chars): Extracts the left part of a text string.
                14) RIGHT(text, num_chars): Extracts the right part of a text string.
                15) MID(text, start_num, num_chars): Extracts a substring from the middle of a text string.
                16) LEN(text): Returns the length of a text string.
                17) LOWER(text): Converts text to lowercase.
                18) UPPER(text): Converts text to uppercase.
                19) TRIM(text): Removes extra spaces from text.
                20) FIND(find_text, within_text, [start_num]): Finds the position of a text within another text (case-sensitive).
                21) SEARCH(find_text, within_text, [start_num]): Finds the position of a text within another text (case-insensitive).
                22) REPLACE(old_text, start_num, num_chars, new_text): Replaces part of a text with another text.
                23) SUBSTITUTE(text, old_text, new_text, [instance_num]): Replaces specific text within a string.
                24) TEXT(value, format_text): Converts a value to text with a specified format.
                25) VALUE(text): Converts text to a numeric value.
                26) PROPER(text): Converts text to proper case (first letter of each word capitalized).
                27) REPT(text, number_times): Repeats text a specified number of times.
                28) EXACT(text1, text2): Compares two text strings to check if they are exactly the same (case-sensitive).
                29) CHAR(number): Returns the character corresponding to an ASCII code.
                30) CODE(text): Returns the ASCII code of the first character in a text string.""")
        elif choice == 2:
            while True:
                try:
                    # Reset quadruples for new expression
                    quad_gen.quads = []
                    
                    expr = input("Enter an Excel expression: ").strip()
                    if expr.startswith("="):
                        expr = expr[1:]
                    result = parser.parse(expr)
                    print("\nResult:", result)
                    
                    # Display generated quadruples
                    quad_gen.print_quads()
                except EOFError:
                    break
        elif choice == 3:
            print("Exiting...")
            break
        else:
            print("Invalid choice. Please try again.")


